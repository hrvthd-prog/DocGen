#!/usr/bin/env python3
"""Elba Sichtreiter papírcetli-sablon generátor (58 x 18 mm, A4, álló).

Kimenet: Sichtreiter-cimkek-58x18mm.xlsx — két lap:
  "Nevek"  : ide illeszti be a felhasználó a névlistát (A2:A84, 83 sor).
  "Címkék" : nyomtatható rács, a nevek formulával jönnek át a Nevek lapról.

Miért char-egység a szélesség? Az OOXML a kolumnaszélességet NEM mm-ben tárolja,
hanem a munkafüzet alap-betűtípusának „maximum digit width" (MDW) egységében.
Az átszámítás alkalmazásfüggő — ezt a konténerben LibreOffice-szal ki is mértük:

    Excel:       képpont = char * MDW_egész + 5      (Calibri 11 -> MDW = 7)
    LibreOffice: képpont = char * MDW_valós          (Calibri 11 -> 7,392)

A kettő szerkezetileg (az +5 tapadás miatt) nem hozható egyszerre pontosra, ezért
a cél explicit. Alapértelmezés: `excel` — ez a dokumentált eset, és a közismert
„8,43 char = 64 képpont" alappal hitelesített (8,43*7+5 = 64,01). A sormagasság
pontban megy, ott nincs ilyen áttétel és nincs alkalmazásfüggés:
pt = mm / 25.4 * 72.

Használat:
    python3 generate.py                      # Excel-pontos (alapértelmezés)
    python3 generate.py --target libreoffice  # LibreOffice/Calc-pontos
"""

import argparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

# --- A megrendelt geometria -------------------------------------------------
LABEL_W_MM = 58.0          # névcella szélesség
LABEL_H_MM = 18.0          # névcella magasság
GAP_MM = 2.0               # vágási rés vízszintesen és függőlegesen
NAMES = 83                 # ennyi névnek kell hely

# --- Lapkiosztás A4 állóra --------------------------------------------------
# Vízszintes: 3 * 58 + 2 * 2 = 178 mm  ->  (210 - 178) / 2 = 16 mm oldalmargó.
# Függőleges: 13 * 18 + 12 * 2 = 258 mm  ->  bőven belül van a 297 mm-en, a
# margó 13 mm, ami minden asztali nyomtató nyomtatható területén belül van.
# (14 sor is kijönne matematikailag — 278 mm —, de akkor a margó 9,5 mm-re
# szűkül, és egyes tintasugarasok alsó nem-nyomtatható sávja ezt levágná.)
COLS = 3
ROWS_PER_PAGE = 13
LABEL_ROWS = -(-NAMES // COLS)            # 28 sor -> 84 hely (83 név + 1 tartalék)

PAGE_W_MM, PAGE_H_MM = 210.0, 297.0
GRID_W_MM = COLS * LABEL_W_MM + (COLS - 1) * GAP_MM

# A margó SZÁNDÉKOSAN kisebb, mint a (210 - 178) / 2 = 16 mm „pontos" érték.
# Ha a nyomtatható sáv épp csak annyi, mint a rács, a legkisebb kerekítési
# eltérés is új lapra tolja a 3. oszlopot (LibreOffice-szal igazolva: 16 mm-es
# margóval az E oszlop külön lapokra került). 10 mm margó -> 190 mm nyomtatható
# sáv 178 mm tartalomra, azaz 12 mm tartalék; a vízszintes középre igazítás
# (horizontalCentered) pedig a papíron visszaadja a szimmetrikus 16 mm-t.
SIDE_MARGIN_MM = 10.0
TOPBOT_MARGIN_MM = 13.0

MM_PER_INCH = 25.4
DPI = 96.0                 # a táblázatkezelők képpont-rácsa (96 DPI)

# A munkafüzet ALAP (Normal) betűtípusa Calibri 11 marad — erre épül a
# mm-pontosság, és ez az openpyxl alapértéke is. A látható szöveg a celláknál
# Arial; a cellaszintű betűtípus a kolumnaszélességet nem érinti.
MDW_EXCEL = 7.0            # Calibri 11 egész MDW (dokumentált; 8,43*7+5 = 64 px)
MDW_LIBRE = 7.392          # Calibri 11 valós MDW — a konténerben kimérve

TARGET = "excel"           # a main() írja át --target szerint


def col_width(mm: float) -> float:
    """mm -> kolumnaszélesség karakter-egységben, 2 tizedesre kerekítve."""
    px = mm / MM_PER_INCH * DPI
    if TARGET == "libreoffice":
        return round(px / MDW_LIBRE, 2)
    return round((px - 5.0) / MDW_EXCEL, 2)


def row_height(mm: float) -> float:
    """mm -> sormagasság pontban."""
    return round(mm / MM_PER_INCH * 72.0, 2)


def inch(mm: float) -> float:
    return round(mm / MM_PER_INCH, 4)


# --- Stílusok ---------------------------------------------------------------
DASHED_GRAY = Side(style="dashed", color="FF808080")
LABEL_BORDER = Border(left=DASHED_GRAY, right=DASHED_GRAY,
                      top=DASHED_GRAY, bottom=DASHED_GRAY)
LABEL_ALIGN = Alignment(horizontal="center", vertical="center",
                        shrink_to_fit=True)
LABEL_FONT = Font(name="Arial", size=12)

INPUT_FILL = PatternFill("solid", fgColor="FFFFF2CC")   # halvány sárga = ide írj
THIN_GRAY = Side(style="thin", color="FFBFBFBF")
INPUT_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY,
                      top=THIN_GRAY, bottom=THIN_GRAY)


def build_names_sheet(ws) -> None:
    """A beviteli lap: A2:A84 a beillesztési sáv, mellette a használati leírás."""
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 2
    ws.column_dimensions["C"].width = 76

    ws["A1"] = f"Nevek (1–{NAMES})"
    ws["A1"].font = Font(name="Arial", size=11, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    for i in range(NAMES):
        cell = ws.cell(row=2 + i, column=1)
        cell.fill = INPUT_FILL
        cell.border = INPUT_BORDER
        cell.font = Font(name="Arial", size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    lines = [
        ("Használat", True),
        (f"1. Illeszd be a névlistát a sárga sávba: A2:A{NAMES + 1} "
         f"({NAMES} sor, egy név egy sorban).", False),
        ("2. Menj a „Címkék” lapra — a nevek automatikusan megjelennek a "
         "rácsban, balról jobbra, fentről le.", False),
        ("3. Nyomtatás (Ctrl+P): A4, álló, méretezés 100% / „Tényleges méret”. "
         "SEMMILYEN „laphoz igazítás” ne legyen bekapcsolva, mert az elrontja "
         "a mm-pontos méretet.", False),
        ("4. Vágás: a szaggatott szürke keretek között 2 mm rés van — a rés "
         "közepén vágj.", False),
        ("5. ELLENŐRZÉS az első lap kinyomtatása után: mérd meg vonalzóval egy "
         f"keret szélességét — {LABEL_W_MM:.0f} mm-nek kell lennie. Ha nem az, "
         "a nyomtatási méretezés nem 100%.", False),
        ("", False),
        ("Példa a várt formátumra (ilyen alakban illeszd be):", True),
        ("Kovácsné dr. Szabó Annamária", False),
        ("", False),
        (f"Címkeméret: {LABEL_W_MM:.0f} × {LABEL_H_MM:.0f} mm · rés: "
         f"{GAP_MM:.0f} mm · {COLS} oszlop × {ROWS_PER_PAGE} sor laponként.", False),
        (f"A rácsban {LABEL_ROWS * COLS} hely van ({NAMES} névhez + "
         f"{LABEL_ROWS * COLS - NAMES} tartalék), 3 A4-es lapon.", False),
        ("", False),
        (f"Megjegyzés a méretpontosságról: ez a fájl "
         f"{'EXCELRE' if TARGET == 'excel' else 'LIBREOFFICE CALC-RA'} van "
         f"kalibrálva (oszlopszélesség {col_width(LABEL_W_MM)} / "
         f"{col_width(GAP_MM)} karakter-egység). A két program másképp számítja "
         f"a karakter-egységet, ezért a másikban a címke kb. "
         f"{'59,9' if TARGET == 'excel' else '55,7'} mm-re jön ki — ahhoz a "
         f"generate.py --target "
         f"{'libreoffice' if TARGET == 'excel' else 'excel'} változat kell.", False),
    ]
    for offset, (text, bold) in enumerate(lines):
        cell = ws.cell(row=1 + offset, column=3, value=text or None)
        cell.font = Font(name="Arial", size=11, bold=bold)
        cell.alignment = Alignment(horizontal="left", vertical="top")

    row = 1 + len(lines) + 1
    ws.cell(row=row, column=3, value="Kitöltött nevek:").font = Font(
        name="Arial", size=11, bold=True)
    count = ws.cell(row=row, column=4, value=f"=COUNTA(A2:A{NAMES + 1})")
    count.font = Font(name="Arial", size=11, bold=True)
    count.alignment = Alignment(horizontal="left")

    ws.freeze_panes = "A2"

    # Ha valaki tévedésből az egész munkafüzetet nyomtatja, ez a lap ne
    # tördelődjön szét: csak a névsáv a nyomtatási terület.
    ws.print_area = f"A1:A{NAMES + 1}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4


def build_labels_sheet(ws) -> None:
    """A nyomtatható rács: 58x18 mm-es cellák, közöttük 2 mm-es üres sáv."""
    # Oszlopok: névcella, rés, névcella, rés, névcella  (A, B, C, D, E)
    name_cols = []
    col = 1
    for i in range(COLS):
        if i:
            ws.column_dimensions[get_column_letter(col)].width = col_width(GAP_MM)
            col += 1
        ws.column_dimensions[get_column_letter(col)].width = col_width(LABEL_W_MM)
        name_cols.append(col)
        col += 1
    last_col = col - 1

    # Sorok: névsor, réssor, névsor, réssor, ... (a névsorok a páratlanok)
    total_rows = LABEL_ROWS * 2 - 1
    for r in range(1, total_rows + 1):
        is_label = r % 2 == 1
        ws.row_dimensions[r].height = row_height(
            LABEL_H_MM if is_label else GAP_MM)

    for n in range(LABEL_ROWS):                 # 0-alapú névsor-index
        r = n * 2 + 1
        for c, col_idx in enumerate(name_cols):
            slot = n * COLS + c + 1             # 1..84, olvasási sorrendben
            cell = ws.cell(row=r, column=col_idx)
            if slot <= NAMES:
                src = f"$A${slot + 1}"          # A1 a fejléc, ezért +1
                cell.value = f'=IF(Nevek!{src}="","",Nevek!{src})'
            cell.border = LABEL_BORDER
            cell.alignment = LABEL_ALIGN
            cell.font = LABEL_FONT

    # Oldaltörések a réssorok után, hogy minden lap névsorral kezdődjön.
    for page_end in range(ROWS_PER_PAGE, LABEL_ROWS, ROWS_PER_PAGE):
        ws.row_breaks.append(Break(id=page_end * 2))   # a réssor után

    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:{get_column_letter(last_col)}{total_rows}"

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.scale = 100
    ws.page_setup.fitToPage = False
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    ws.page_margins.left = inch(SIDE_MARGIN_MM)
    ws.page_margins.right = inch(SIDE_MARGIN_MM)
    ws.page_margins.top = inch(TOPBOT_MARGIN_MM)
    ws.page_margins.bottom = inch(TOPBOT_MARGIN_MM)
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0
    ws.oddHeader.center.text = None
    ws.oddFooter.center.text = None


def main() -> None:
    global TARGET
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--target", choices=("excel", "libreoffice"), default="excel",
                    help="melyik alkalmazásban legyen mm-pontos a szélesség")
    ap.add_argument("-o", "--out", default=None, help="kimeneti fájlnév")
    args = ap.parse_args()
    TARGET = args.target

    wb = Workbook()                             # alap (Normal) betűtípus: Calibri 11
    names = wb.active
    names.title = "Nevek"
    labels = wb.create_sheet("Címkék")

    build_names_sheet(names)
    build_labels_sheet(labels)

    wb.active = wb.index(labels)
    out = args.out or ("Sichtreiter-cimkek-58x18mm.xlsx" if TARGET == "excel"
                       else "Sichtreiter-cimkek-58x18mm-libreoffice.xlsx")
    wb.save(out)

    print(f"{out} kész  (cél: {TARGET})")
    print(f"  névcella  : {LABEL_W_MM} x {LABEL_H_MM} mm "
          f"-> szélesség {col_width(LABEL_W_MM)} char, "
          f"magasság {row_height(LABEL_H_MM)} pt")
    print(f"  rés       : {GAP_MM} mm -> szélesség {col_width(GAP_MM)} char, "
          f"magasság {row_height(GAP_MM)} pt")
    print(f"  rács      : {COLS} x {LABEL_ROWS} = {COLS * LABEL_ROWS} hely, "
          f"{GRID_W_MM} mm széles")
    print(f"  margó     : oldal {SIDE_MARGIN_MM} mm, "
          f"fent/lent {TOPBOT_MARGIN_MM} mm")


if __name__ == "__main__":
    main()
